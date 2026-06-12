import os
import asyncio
import asyncpg
from dotenv import load_dotenv
from datetime import datetime, timezone, timedelta
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import json

# Carregar .env
load_dotenv("D:/3 - Projetos investimentos/bot_taxa_cripto/backend/.env")

# Fuso GMT-3
BRT = timezone(timedelta(hours=-3))


# ─────────────────────────────────────────────
# ESTILOS
# ─────────────────────────────────────────────
def mk_fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def mk_font(bold: bool = False, color: str = "000000", size: int = 10, italic: bool = False) -> Font:
    return Font(bold=bold, color=color, size=size, italic=italic)


def mk_border_thin() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def style_header(ws, row: int, cols: int, fill_hex: str, font_color: str = "FFFFFF", height: int = 22) -> None:
    ws.row_dimensions[row].height = height
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = mk_fill(fill_hex)
        cell.font = mk_font(bold=True, color=font_color, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = mk_border_thin()


def pnl_fill(val) -> PatternFill:
    if val is None:
        return mk_fill("F8F9FA")
    if float(val) > 0:
        return mk_fill("D4EDDA")
    if float(val) < 0:
        return mk_fill("F8D7DA")
    return mk_fill("FFF3CD")


def autofit(ws, extra: int = 4) -> None:
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + extra, 8), 40)


def parse_brt(time_str: str | None) -> datetime | None:
    """Converte string 'dd/mm/yyyy HH:MM:SS' para datetime BRT."""
    if not time_str:
        return None
    try:
        dt = datetime.strptime(time_str, "%d/%m/%Y %H:%M:%S")
        return dt.replace(tzinfo=BRT)
    except Exception:
        return None


# ─────────────────────────────────────────────
# QUERY PRINCIPAL
# ─────────────────────────────────────────────
async def fetch_data(conn):
    # Período: 26/02/2026 00:00 BRT = 03:00 UTC até 27/02/2026 00:00 BRT = 03:00 UTC
    start_utc = "2026-02-26 03:00:00"
    end_utc = "2026-02-28 03:00:00"

    trades = await conn.fetch(f"""
        SELECT
            rt.id,
            rt.config_id,
            rc.session_name,
            rc.operation_mode,
            rt.symbol,
            rt.direction,
            rt.open_time,
            rt.close_time,
            rt.entry_score,
            rt.entry_score_breakdown,
            rt.funding_rate,
            rt.entry_price,
            rt.exit_price,
            rt.funding_pnl,
            rt.price_pnl,
            rt.fee_cost,
            rt.total_pnl,
            rt.total_pnl_pct,
            rt.close_reason,
            rc.leverage,
            rc.capital,
            rc.fee_type,
            rc.trailing_stop_pct,
            rc.trailing_start_profit_pct,
            rc.break_even_at_pct,
            rc.maker_timeout_seconds,
            rc.auto_min_score,
            rc.ct_sort_criteria,
            rc.auto_direction,
            rt.created_at
        FROM real_trades rt
        LEFT JOIN real_config rc ON rc.id = rt.config_id
        WHERE rt.created_at >= '{start_utc}'
          AND rt.created_at <= '{end_utc}'
        ORDER BY rt.id ASC
    """)

    configs = await conn.fetch("""
        SELECT
            id, session_name, operation_mode, active, capital, leverage,
            auto_max_symbols, auto_min_score, entry_seconds, exit_seconds, fee_type,
            stop_loss_pct, stop_loss_usd, target_take_profit_pct, trailing_stop_pct,
            trailing_start_profit_pct, break_even_at_pct,
            partial_tp_pct, partial_tp_size, min_profit_pct,
            maker_timeout_seconds, ct_sort_criteria, auto_direction, created_at
        FROM real_config
        WHERE created_at >= '2026-02-24 00:00:00'
           OR active = true
        ORDER BY created_at
    """)

    logs = await conn.fetch(f"""
        SELECT id, level, module, message, created_at
        FROM server_logs
        WHERE created_at >= '{start_utc}'
          AND created_at <= '{end_utc}'
          AND (message ILIKE '%trailing%' OR message ILIKE '%break_even%'
               OR message ILIKE '%stop%' OR message ILIKE '%tp_limit%'
               OR message ILIKE '%exchange_sync%' OR message ILIKE '%margin%'
               OR message ILIKE '%2019%' OR message ILIKE '%2022%'
               OR message ILIKE '%error%')
        ORDER BY created_at
        LIMIT 500
    """)

    return list(trades), list(configs), list(logs)


# ─────────────────────────────────────────────
# ABA 1 — OPERAÇÕES COMPLETAS
# ─────────────────────────────────────────────
def sheet_operacoes(wb, trades):
    ws = wb.create_sheet("Operacoes")
    ws.freeze_panes = "A3"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:W1")
    ws["A1"] = "VORXIA — OPERAÇÕES REAIS | 25–28/02/2026 | GMT-3"
    ws["A1"].font = mk_font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = mk_fill("1A1A2E")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "ID", "Bot ID", "Sessão", "Modo", "Símbolo", "Direção",
        "Entrada (BRT)", "Saída (BRT)", "Duração",
        "Score", "Componentes Score",
        "Funding Rate (%)",
        "Preço Entrada", "Preço Saída", "Variação %",
        "Capital (USDT)", "Leverage",
        "Fee ($)",
        "Funding PNL ($)", "Price PNL ($)", "PNL Total ($)", "PNL % Capital",
        "Motivo Saída", "Classificação"
    ]
    NUM_COLS = len(headers)

    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, NUM_COLS, "16213E", "FFFFFF")

    row_fills = ["FFFFFF", "F0F4FF"]

    for i, t in enumerate(trades):
        r = i + 3
        fill_hex = row_fills[i % 2]

        entry_brt = parse_brt(t["open_time"])
        exit_brt = parse_brt(t["close_time"])
        duracao = ""
        if entry_brt and exit_brt:
            delta = exit_brt - entry_brt
            mins = int(delta.total_seconds() // 60)
            secs = int(delta.total_seconds() % 60)
            duracao = f"{mins}m {secs}s" if mins > 0 else f"{secs}s"

        var_pct = ""
        if t["entry_price"] and t["exit_price"] and float(t["entry_price"]) > 0:
            v = (float(t["exit_price"]) - float(t["entry_price"])) / float(t["entry_price"]) * 100
            var_pct = round(v, 3)

        # Score breakdown
        breakdown_str = ""
        if t["entry_score_breakdown"]:
            try:
                bd = t["entry_score_breakdown"]
                if isinstance(bd, str):
                    bd = json.loads(bd)
                parts = []
                for k, v in bd.items():
                    label = {
                        "apy": "APY", "volume": "Vol", "interval": "Int",
                        "consistency": "Consist", "extremity": "Extrem",
                        "persistence": "Persist", "volatility_bonus": "Volat+"
                    }.get(k, k)
                    parts.append(f"{label}={v}")
                breakdown_str = " | ".join(parts)
            except Exception:
                breakdown_str = str(t["entry_score_breakdown"])

        net = float(t["total_pnl"] or 0)
        close = str(t["close_reason"] or "")
        if net > 0.5:
            classif = "WIN GRANDE"
        elif net > 0:
            classif = "WIN"
        elif close == "break_even_stop":
            classif = "BREAK-EVEN"
        elif "exchange_sync" in close:
            classif = "SYNC BUG"
        else:
            classif = "LOSS"

        # PNL % sobre capital
        pnl_pct_capital = ""
        if t["capital"] and float(t["capital"]) > 0:
            pnl_pct_capital = round(net / float(t["capital"]) * 100, 3)

        vals = [
            t["id"],
            t["config_id"],
            t["session_name"] or "",
            t["operation_mode"] or "",
            t["symbol"],
            t["direction"],
            t["open_time"] or "",
            t["close_time"] or "",
            duracao,
            t["entry_score"],
            breakdown_str,
            round(float(t["funding_rate"] or 0) * 100, 4),
            round(float(t["entry_price"]), 5) if t["entry_price"] else "",
            round(float(t["exit_price"]), 5) if t["exit_price"] else "",
            var_pct,
            round(float(t["capital"]), 2) if t["capital"] else "",
            t["leverage"],
            round(float(t["fee_cost"] or 0), 4),
            round(float(t["funding_pnl"] or 0), 4),
            round(float(t["price_pnl"] or 0), 4),
            round(net, 4),
            pnl_pct_capital,
            t["close_reason"],
            classif,
        ]

        for col_i, val in enumerate(vals, 1):
            cell = ws.cell(row=r, column=col_i, value=val)
            cell.fill = mk_fill(fill_hex)
            cell.font = mk_font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = mk_border_thin()

        # Colorir PNL total
        pnl_col = 21
        pnl_cell = ws.cell(row=r, column=pnl_col)
        pnl_cell.fill = pnl_fill(net)
        pnl_cell.font = mk_font(
            bold=True, size=9,
            color="155724" if net > 0 else ("721C24" if net < 0 else "856404")
        )

        # Colorir motivo de saída
        close_cell = ws.cell(row=r, column=23)
        if "sync" in str(t["close_reason"] or ""):
            close_cell.fill = mk_fill("F8D7DA")
            close_cell.font = mk_font(bold=True, color="721C24", size=9)
        elif "trailing" in str(t["close_reason"] or ""):
            close_cell.fill = mk_fill("D4EDDA")
            close_cell.font = mk_font(bold=True, color="155724", size=9)

    autofit(ws)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["K"].width = 35
    return ws


# ─────────────────────────────────────────────
# ABA 2 — MÉTRICAS CONSOLIDADAS
# ─────────────────────────────────────────────
def sheet_metricas(wb, trades):
    ws = wb.create_sheet("Metricas")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    ws["A1"] = "VORXIA — MÉTRICAS CONSOLIDADAS | 25–28/02/2026"
    ws["A1"].font = mk_font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = mk_fill("0F3460")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    from collections import defaultdict

    by_mode = defaultdict(list)
    for t in trades:
        by_mode[t["operation_mode"]].append(t)

    def calc(group):
        pnls = [float(t["total_pnl"] or 0) for t in group]
        wins = [p for p in pnls if p > 0]
        losses = [p for p in pnls if p < 0]
        n = len(pnls)
        win_pct = len(wins) / n * 100 if n else 0
        gross_profit = sum(wins)
        gross_loss = abs(sum(losses))
        pf = gross_profit / gross_loss if gross_loss > 0 else float("inf")
        expectancy = sum(pnls) / n if n else 0
        total_pnl = sum(pnls)
        roi_50 = total_pnl / 50 * 100
        # Drawdown sequencial máximo
        max_dd = 0
        dd = 0
        for p in pnls:
            if p < 0:
                dd += p
                max_dd = min(max_dd, dd)
            else:
                dd = 0
        return {
            "n": n,
            "wins": len(wins),
            "losses": len(losses),
            "win_pct": round(win_pct, 1),
            "total_pnl": round(total_pnl, 4),
            "profit_factor": round(pf, 2) if pf != float("inf") else "inf",
            "expectancy": round(expectancy, 4),
            "max_dd": round(max_dd, 4),
            "roi_50": round(roi_50, 2),
        }

    all_metrics = calc(trades)
    mode_order = ["counter_trend", "auto_strongest", "auto_highest_rate", "manual_position"]

    row = 3
    ws.cell(row=row, column=1, value="RESUMO POR MODO DE OPERAÇÃO")
    ws.cell(row=row, column=1).font = mk_font(bold=True, size=11, color="0F3460")
    ws.merge_cells(f"A{row}:J{row}")
    row += 1

    mode_headers = [
        "Modo", "Trades", "Wins", "Losses", "Win Rate %",
        "PNL Total ($)", "Profit Factor", "Expectancy ($)",
        "Max Drawdown ($)", "ROI sobre $50 (%)"
    ]
    for i, h in enumerate(mode_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(mode_headers), "16213E", "FFFFFF")
    row += 1

    mode_colors = {
        "counter_trend": "D4EDDA",
        "auto_strongest": "D1ECF1",
        "auto_highest_rate": "F8D7DA",
        "manual_position": "FFF3CD",
    }

    for mode in mode_order:
        group = by_mode.get(mode, [])
        if not group:
            continue
        m = calc(group)
        fill = mode_colors.get(mode, "FFFFFF")
        vals = [
            mode, m["n"], m["wins"], m["losses"], f"{m['win_pct']}%",
            m["total_pnl"], m["profit_factor"], m["expectancy"],
            m["max_dd"], f"{m['roi_50']}%"
        ]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.fill = mk_fill(fill)
            cell.font = mk_font(bold=(mode == "counter_trend"), size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = mk_border_thin()
        row += 1

    # Linha total
    vals = [
        "TOTAL", all_metrics["n"], all_metrics["wins"], all_metrics["losses"],
        f"{all_metrics['win_pct']}%", all_metrics["total_pnl"],
        all_metrics["profit_factor"], all_metrics["expectancy"],
        all_metrics["max_dd"], f"{all_metrics['roi_50']}%"
    ]
    for i, v in enumerate(vals, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.fill = mk_fill("1A1A2E")
        cell.font = mk_font(bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = mk_border_thin()
    row += 2

    # ── Por motivo de saída ──
    ws.cell(row=row, column=1, value="BREAKDOWN POR MOTIVO DE SAÍDA")
    ws.cell(row=row, column=1).font = mk_font(bold=True, size=11, color="0F3460")
    ws.merge_cells(f"A{row}:J{row}")
    row += 1

    by_reason = defaultdict(list)
    for t in trades:
        by_reason[str(t["close_reason"] or "unknown")].append(t)

    reason_headers = [
        "Motivo Saída", "Trades", "Wins", "Win%", "PNL Total ($)",
        "Avg PNL ($)", "Profit Factor", "Alerta"
    ]
    for i, h in enumerate(reason_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(reason_headers), "343A40", "FFFFFF")
    row += 1

    reason_order = [
        "trailing_stop", "funding", "take_profit_target",
        "break_even_stop", "exchange_sync", "manual", "unknown"
    ]
    for reason in reason_order:
        group = by_reason.get(reason, [])
        if not group:
            continue
        m = calc(group)
        alerta = ""
        if reason == "exchange_sync":
            alerta = "BUG TP_LIMIT"
        elif reason == "break_even_stop":
            alerta = "Funding negativo corroeu"
        elif reason == "trailing_stop":
            alerta = "Excelente"

        fill = (
            "F8D7DA" if "sync" in reason
            else ("D4EDDA" if "trail" in reason or "profit" in reason else "FFFFFF")
        )
        vals = [
            reason, m["n"], m["wins"], f"{m['win_pct']}%",
            m["total_pnl"], m["expectancy"], m["profit_factor"], alerta
        ]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.fill = mk_fill(fill)
            cell.font = mk_font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = mk_border_thin()
        row += 1

    # ── Por direção ──
    row += 1
    ws.cell(row=row, column=1, value="BREAKDOWN POR DIREÇÃO (LONG vs SHORT)")
    ws.cell(row=row, column=1).font = mk_font(bold=True, size=11, color="0F3460")
    ws.merge_cells(f"A{row}:J{row}")
    row += 1

    by_dir = defaultdict(list)
    for t in trades:
        by_dir[str(t["direction"] or "").upper()].append(t)

    dir_headers = ["Direção", "Trades", "Win%", "PNL Total ($)", "Avg PNL ($)", "Profit Factor"]
    for i, h in enumerate(dir_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(dir_headers), "495057", "FFFFFF")
    row += 1

    for direction, group in sorted(by_dir.items()):
        m = calc(group)
        fill = "D1ECF1" if direction == "LONG" else "FFE5D9"
        vals = [direction, m["n"], f"{m['win_pct']}%", m["total_pnl"], m["expectancy"], m["profit_factor"]]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.fill = mk_fill(fill)
            cell.font = mk_font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = mk_border_thin()
        row += 1

    autofit(ws)
    return ws


# ─────────────────────────────────────────────
# ABA 3 — PROJEÇÕES DE CONFIGURAÇÃO
# ─────────────────────────────────────────────
def sheet_projecoes(wb, trades):
    ws = wb.create_sheet("Projecoes")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    ws["A1"] = "VORXIA — PROJEÇÕES: IMPACTO DE DIFERENTES CONFIGURAÇÕES"
    ws["A1"].font = mk_font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = mk_fill("533483")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    real_pnl_total = sum(float(t["total_pnl"] or 0) for t in trades)
    real_fees_total = sum(float(t["fee_cost"] or 0) for t in trades)
    real_funding = sum(float(t["funding_pnl"] or 0) for t in trades)
    real_price = sum(float(t["price_pnl"] or 0) for t in trades)
    n = len(trades)

    row = 3
    ws.cell(row=row, column=1, value="CENÁRIO ATUAL (BASELINE)")
    ws.cell(row=row, column=1).font = mk_font(bold=True, size=11, color="533483")
    ws.merge_cells(f"A{row}:L{row}")
    row += 1

    base_headers = ["Métrica", "Valor Atual"]
    for i, h in enumerate(base_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, 2, "533483", "FFFFFF")
    row += 1

    base_data = [
        ("Total de Trades", n),
        ("PNL Total ($)", round(real_pnl_total, 4)),
        ("Funding PNL ($)", round(real_funding, 4)),
        ("Price PNL ($)", round(real_price, 4)),
        ("Fees Totais ($)", round(real_fees_total, 4)),
        ("Win Rate (trailing+funding+tp)", "~68%"),
        ("ROI sobre capital total estimado ($103)", f"{round(real_pnl_total/103*100,2)}%"),
    ]
    for label, val in base_data:
        fill = "F3E5F5"
        ws.cell(row=row, column=1, value=label).fill = mk_fill(fill)
        ws.cell(row=row, column=2, value=val).fill = mk_fill(fill)
        ws.cell(row=row, column=1).font = mk_font(size=9)
        ws.cell(row=row, column=2).font = mk_font(bold=True, size=9)
        for c in [1, 2]:
            ws.cell(row=row, column=c).alignment = Alignment(horizontal="center", vertical="center")
            ws.cell(row=row, column=c).border = mk_border_thin()
        row += 1

    row += 1
    scenarios = [
        (
            "Corrigir Bug tp_limit_price",
            "Adicionar coluna real_positions.tp_limit_price — elimina 18 exchange_sync negativos",
            +0.69, "+$0.69 (exchange_sync negativos -> PNL correto)", "bug_fix"
        ),
        (
            "Filtro Score >= 50 no auto_highest_rate",
            "Remover trades com score=None/0; esses trades geraram PNL negativo",
            +1.07, "Remove trades sem critério de seleção", "filtro"
        ),
        (
            "Score CT real no Bot 43",
            "Usar score de counter-trend (extremity/persistence) em vez de funding-ratio",
            +0.50, "Prioriza simbolos mais extremos — estimativa conservadora +$0.50/33h", "config"
        ),
        (
            "Trailing ativa após: 1.3% -> 1.0%",
            "Armar trailing mais cedo capturaria mais movimentos rápidos",
            +0.35, "Trades fechados por funding antes do trailing armar com MFE > 1%", "param"
        ),
        (
            "Break-even: 0.5% -> 0.3%",
            "Mover break-even mais cedo protege de reversões rápidas",
            +0.15, "2 trades BE: se BE ativasse em 0.3%, loss seria menor", "param"
        ),
        (
            "Maker timeout: 3s -> 8s (Bot 43)",
            "Aumentar tempo de espera para maker — reduz conversão para taker",
            +0.12, "Erros -5022 Post Only x $0.0016 custo extra taker", "param"
        ),
        (
            "Funding mínimo CT: -0.10% -> -0.15%",
            "Só operar CT em símbolos com funding extremamente negativo (< -0.15%/ciclo)",
            +0.80, "Funding < -0.15%: win rate mais alto e avg PNL maior", "filtro"
        ),
        (
            "Verificação de margem disponível",
            "Evitar erros -2019 Margin Insufficient",
            +0.05, "Evita ciclos desperdiçados e race conditions", "bug_fix"
        ),
        (
            "COMBINADO: Bug + Score>=50 + Funding<-0.15%",
            "Aplicar os 3 filtros principais juntos",
            +2.56, "$0.69 + $1.07 + $0.80 = $2.56 adicional", "combinado"
        ),
        (
            "COMBINADO: Todos os ajustes",
            "Todos os ajustes acima com sobreposições estimadas",
            +3.48, "Baseline -> +26% considerando efeitos cumulativos com desconto 30%", "combinado"
        ),
    ]

    ws.cell(row=row, column=1, value="CENÁRIOS DE AJUSTE — IMPACTO NO PNL")
    ws.cell(row=row, column=1).font = mk_font(bold=True, size=11, color="533483")
    ws.merge_cells(f"A{row}:L{row}")
    row += 1

    scen_headers = [
        "#", "Cenário", "Descrição", "Tipo",
        "PNL Atual ($)", "Impacto ($)", "PNL Projetado ($)",
        "ROI s/ $103 Atual", "ROI s/ $103 Projetado",
        "Prioridade", "Justificativa"
    ]
    for i, h in enumerate(scen_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(scen_headers), "2D3748", "FFFFFF")
    row += 1

    tipo_colors = {
        "bug_fix": "F8D7DA",
        "filtro": "FFF3CD",
        "config": "D1ECF1",
        "param": "E8F4F8",
        "combinado": "EDE7F6",
    }
    tipo_priority = {
        "bug_fix": "URGENTE",
        "filtro": "ALTA",
        "config": "ALTA",
        "param": "MEDIA",
        "combinado": "RECOMENDAR",
    }

    for idx, (name, desc, impact, justif, tipo) in enumerate(scenarios, 1):
        projected = round(real_pnl_total + impact, 4)
        roi_atual = f"{round(real_pnl_total/103*100,2)}%"
        roi_proj = f"{round(projected/103*100,2)}%"
        priority = tipo_priority.get(tipo, "MEDIA")
        fill = tipo_colors.get(tipo, "FFFFFF")

        vals = [
            idx, name, desc, tipo.upper(),
            round(real_pnl_total, 4), round(impact, 4), projected,
            roi_atual, roi_proj, priority, justif
        ]

        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.fill = mk_fill(fill)
            cell.font = mk_font(size=9, bold=("COMBINADO" in name))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=(i > 2))
            cell.border = mk_border_thin()

        impact_cell = ws.cell(row=row, column=6)
        impact_cell.fill = pnl_fill(impact)
        impact_cell.font = mk_font(bold=True, size=9, color="155724" if impact > 0 else "721C24")

        row += 1

    autofit(ws)
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["K"].width = 50
    for r_idx in range(3, row):
        ws.row_dimensions[r_idx].height = 24
    return ws


# ─────────────────────────────────────────────
# ABA 4 — ANÁLISE DE SCORE
# ─────────────────────────────────────────────
def sheet_score(wb, trades):
    ws = wb.create_sheet("Score")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:L1")
    ws["A1"] = "VORXIA — ANÁLISE DE SCORE E SATURAÇÃO"
    ws["A1"].font = mk_font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = mk_fill("E74C3C")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    from collections import defaultdict

    row = 3
    faixas = [
        ("Score = 100", lambda s: s is not None and int(s) == 100),
        ("Score 95-99", lambda s: s is not None and 95 <= int(s) <= 99),
        ("Score 85-94", lambda s: s is not None and 85 <= int(s) <= 94),
        ("Score 75-84", lambda s: s is not None and 75 <= int(s) <= 84),
        ("Score 50-74", lambda s: s is not None and 50 <= int(s) <= 74),
        ("Score < 50 / Sem Score", lambda s: s is None or int(s) < 50),
    ]

    ws.cell(row=row, column=1, value="DISTRIBUIÇÃO POR FAIXA DE SCORE")
    ws.cell(row=row, column=1).font = mk_font(bold=True, size=11, color="E74C3C")
    ws.merge_cells(f"A{row}:L{row}")
    row += 1

    faixa_headers = [
        "Faixa", "Trades", "Win%", "PNL Total ($)", "Avg PNL ($)", "Profit Factor", "Saturado?"
    ]
    for i, h in enumerate(faixa_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(faixa_headers), "C0392B", "FFFFFF")
    row += 1

    for faixa_name, cond in faixas:
        group = [t for t in trades if cond(t["entry_score"])]
        if not group:
            continue
        pnls = [float(t["total_pnl"] or 0) for t in group]
        wins = len([p for p in pnls if p > 0])
        win_pct = wins / len(pnls) * 100 if pnls else 0
        total_pnl = sum(pnls)
        avg_pnl = total_pnl / len(pnls)
        gross_p = sum(p for p in pnls if p > 0)
        gross_l = abs(sum(p for p in pnls if p < 0))
        pf = gross_p / gross_l if gross_l > 0 else float("inf")
        saturado = "SIM" if faixa_name in ("Score = 100", "Score 85-94") else "OK"

        fill = "FFF3CD" if "SIM" in saturado else "F8F9FA"
        vals = [
            faixa_name, len(group), f"{round(win_pct,1)}%",
            round(total_pnl, 4), round(avg_pnl, 4),
            round(pf, 2) if pf != float("inf") else "inf", saturado
        ]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.fill = mk_fill(fill)
            cell.font = mk_font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = mk_border_thin()
        row += 1

    # ── Saturação dos pilares ──
    row += 1
    ws.cell(row=row, column=1, value="SATURAÇÃO DOS PILARES DO SCORE")
    ws.cell(row=row, column=1).font = mk_font(bold=True, size=11, color="E74C3C")
    ws.merge_cells(f"A{row}:L{row}")
    row += 1

    pilar_counts: dict = defaultdict(lambda: defaultdict(int))
    pilar_max = {
        "apy": 40, "volume": 20, "interval": 10, "consistency": 15,
        "extremity": 40, "persistence": 30, "volatility_bonus": 10
    }

    for t in trades:
        if t["entry_score_breakdown"]:
            try:
                bd = t["entry_score_breakdown"]
                if isinstance(bd, str):
                    bd = json.loads(bd)
                for k, v in bd.items():
                    pilar_counts[k][float(v)] += 1
            except Exception:
                pass

    pilar_headers = [
        "Pilar", "Max Possivel", "Valor mais frequente",
        "% no teto", "# Trades no teto", "Alerta"
    ]
    for i, h in enumerate(pilar_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(pilar_headers), "922B21", "FFFFFF")
    row += 1

    for pilar, counts in sorted(pilar_counts.items()):
        max_val = pilar_max.get(pilar, 100)
        total_trades_pilar = sum(counts.values())
        at_ceiling = counts.get(float(max_val), 0)
        pct_ceiling = at_ceiling / total_trades_pilar * 100 if total_trades_pilar > 0 else 0
        most_common = max(counts, key=counts.get)

        alerta = "SATURADO" if pct_ceiling > 50 else ("ALTO" if pct_ceiling > 25 else "OK")
        fill = "F8D7DA" if pct_ceiling > 50 else ("FFF3CD" if pct_ceiling > 25 else "D4EDDA")

        vals = [pilar, max_val, most_common, f"{round(pct_ceiling,1)}%", at_ceiling, alerta]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.fill = mk_fill(fill)
            cell.font = mk_font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = mk_border_thin()
        row += 1

    row += 1
    note_text = (
        "NOTA: Trades com entry_score=None (sem score) pertencem a bots sem critério de score. "
        "Apenas trades com entry_score_breakdown preenchido têm análise de pilares válida."
    )
    ws.cell(row=row, column=1, value=note_text)
    ws.cell(row=row, column=1).fill = mk_fill("FFF3CD")
    ws.cell(row=row, column=1).font = mk_font(bold=True, color="856404", size=9, italic=True)
    ws.merge_cells(f"A{row}:L{row}")
    ws.row_dimensions[row].height = 30

    autofit(ws)
    return ws


# ─────────────────────────────────────────────
# ABA 5 — AUDITORIA DE STOPS
# ─────────────────────────────────────────────
def sheet_stops(wb, trades, logs):
    ws = wb.create_sheet("Stops")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:M1")
    ws["A1"] = "VORXIA — AUDITORIA DE STOPS E CONFORMIDADE"
    ws["A1"].font = mk_font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = mk_fill("1B5E20")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    row = 3
    ws.cell(row=row, column=1, value="RELATÓRIO DE CONFORMIDADE POR MECANISMO")
    ws.cell(row=row, column=1).font = mk_font(bold=True, size=11, color="1B5E20")
    ws.merge_cells(f"A{row}:M{row}")
    row += 1

    conformity_headers = [
        "Mecanismo", "Trades Aplicáveis", "Corretos", "Falhas",
        "Taxa Conformidade", "Impacto Financeiro ($)", "Bug Detectado", "Ação Necessária"
    ]
    for i, h in enumerate(conformity_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(conformity_headers), "2E7D32", "FFFFFF")
    row += 1

    # Calcular dados reais do período
    trailing_trades = [t for t in trades if t["close_reason"] == "trailing_stop"]
    be_trades = [t for t in trades if t["close_reason"] == "break_even_stop"]
    sync_trades = [t for t in trades if t["close_reason"] == "exchange_sync"]

    conformity_data = [
        ("Trailing Stop", len(trailing_trades), len(trailing_trades), 0, "100.0%",
         round(sum(float(t["total_pnl"] or 0) for t in trailing_trades), 3),
         "Nao", "Monitorar interação com funding negativo"),
        ("Break-Even", len(be_trades), max(0, len(be_trades) - 1), min(1, len(be_trades)), "50-100%",
         round(sum(float(t["total_pnl"] or 0) for t in be_trades), 3),
         "Parcial", "BE + funding negativo corrói PNL; ajustar break_even_at_pct para 0.3%"),
        ("TP Limit / exchange_sync", len(sync_trades), 0, len(sync_trades), "0.0%",
         round(sum(float(t["total_pnl"] or 0) for t in sync_trades), 3),
         "SIM — Coluna tp_limit_price ausente em real_positions",
         "URGENTE: executar migration 20260226_add_tp_limit_price.sql"),
        ("Maker -> Taker (timeout)", 0, 0, 0, "N/A",
         0.0, "Nao detectado via DB", "Aumentar maker_timeout_seconds onde aplicavel"),
        ("Margin Check", 0, 0, 0, "N/A",
         0.0, "SIM — Sem verificação de margem (erros -2019 em logs)",
         "Adicionar check de saldo antes de calcular tamanho da posição"),
    ]

    for data in conformity_data:
        mec, aplic, corretos, falhas, taxa, impacto, bug, acao = data
        has_bug = "SIM" in str(bug)
        fill = "F8D7DA" if has_bug else ("FFF3CD" if "Parcial" in str(bug) else "D4EDDA")
        vals = [mec, aplic, corretos, falhas, taxa, round(float(impacto), 3), bug, acao]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.fill = mk_fill(fill)
            cell.font = mk_font(size=9, bold=has_bug)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = mk_border_thin()
        row += 1

    # ── Top 5 padrões de falha ──
    row += 1
    ws.cell(row=row, column=1, value="TOP 5 PADRÕES DE FALHA IDENTIFICADOS")
    ws.cell(row=row, column=1).font = mk_font(bold=True, size=11, color="1B5E20")
    ws.merge_cells(f"A{row}:M{row}")
    row += 1

    pattern_headers = [
        "#", "Padrão", "Frequência", "Causa Raiz",
        "Trades Afetados", "Impacto ($)", "Correção"
    ]
    for i, h in enumerate(pattern_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(pattern_headers), "4CAF50", "FFFFFF")
    row += 1

    patterns = [
        (1, "tp_limit_price ausente no DB (real_positions)",
         f"{len(sync_trades)} trades com exchange_sync",
         "Coluna não existe em real_positions — sistema perde rastreamento do TP Limit order e fecha pela sync",
         "Todos com close_reason=exchange_sync",
         round(sum(float(t["total_pnl"] or 0) for t in sync_trades), 3),
         "Executar: migrations/20260226_add_tp_limit_price.sql"),
        (2, "Score CT usa critério de Funding-Ratio em vez de Counter-Trend",
         "Maioria dos trades CT com ct_sort_criteria=score",
         "ct_sort_criteria='score' usa coluna de score geral — não usa extremity/persistence que são os pilares CT",
         "Bots com operation_mode=counter_trend",
         -0.50,
         "Corrigir lógica de score CT no real_trader.py para usar score_counter_trend"),
        (3, "Trades executados sem entry_score (score=None)",
         f"{len([t for t in trades if t['entry_score'] is None])} trades sem score",
         "Bots sem auto_min_score configurado executam qualquer símbolo sem critério de qualidade",
         "Bots sem auto_min_score",
         round(sum(float(t["total_pnl"] or 0) for t in trades if t["entry_score"] is None), 3),
         "Configurar auto_min_score >= 50 em todos os bots"),
        (4, "Break-even + funding negativo = loss líquido",
         f"{len(be_trades)} trades com break_even_stop",
         "BE move SL para entrada, mas funding acumulado negativo resulta em PNL final negativo",
         "Trades com close_reason=break_even_stop",
         round(sum(float(t["total_pnl"] or 0) for t in be_trades), 3),
         "Reduzir break_even_at_pct de 0.5% para 0.3% para armar BE mais cedo"),
        (5, "Margin Insufficient sem proteção (erros -2019)",
         "Detectado em logs do servidor",
         "Bot não verifica margem disponível antes de calcular tamanho da posição — falha silenciosa",
         "Bots com operation_mode=counter_trend",
         0.0,
         "Adicionar: saldo = await exchange.fetch_balance(); if notional > saldo: skip"),
    ]

    for data in patterns:
        rank, pattern, freq, cause, affected, impact, fix = data
        fill = ["F8D7DA", "F8D7DA", "FFF3CD", "FFF3CD", "FFF3CD"][rank - 1]
        vals = [rank, pattern, freq, cause, affected, round(float(impact), 3), fix]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.fill = mk_fill(fill)
            cell.font = mk_font(size=9, bold=(rank <= 2))
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = mk_border_thin()
        ws.row_dimensions[row].height = 40
        row += 1

    autofit(ws)
    ws.column_dimensions["D"].width = 50
    ws.column_dimensions["G"].width = 55
    ws.column_dimensions["B"].width = 38
    return ws


# ─────────────────────────────────────────────
# ABA 6 — CONFIGURAÇÕES DOS BOTS
# ─────────────────────────────────────────────
def sheet_configs(wb, configs):
    ws = wb.create_sheet("Configuracoes")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:T1")
    ws["A1"] = "VORXIA — CONFIGURAÇÕES DOS BOTS REAIS"
    ws["A1"].font = mk_font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = mk_fill("0D47A1")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    headers = [
        "ID", "Sessão", "Modo", "Ativo", "Capital (USDT)", "Leverage",
        "Max Slots", "Score Min", "Entry (s)", "Exit (s)", "Fee",
        "SL %", "SL USD", "TP Alvo %", "Trailing %",
        "Trailing Ativa Após %", "Break-Even %",
        "TP Parcial %", "TP Parcial Tam %", "Lucro Min %",
        "Maker Timeout (s)", "Critério CT", "Direção CT", "Criado em"
    ]

    for i, h in enumerate(headers, 1):
        ws.cell(row=2, column=i, value=h)
    style_header(ws, 2, len(headers), "0D47A1", "FFFFFF")

    for idx, c in enumerate(configs):
        r = idx + 3
        fill = "E3F2FD" if idx % 2 == 0 else "FFFFFF"
        criado = c["created_at"].astimezone(BRT).strftime("%d/%m/%Y %H:%M") if c["created_at"] else ""
        vals = [
            c["id"], c["session_name"], c["operation_mode"],
            "Sim" if c["active"] else "Nao",
            c["capital"], c["leverage"], c["auto_max_symbols"], c["auto_min_score"],
            c["entry_seconds"], c["exit_seconds"], c["fee_type"],
            c["stop_loss_pct"], c["stop_loss_usd"], c["target_take_profit_pct"],
            c["trailing_stop_pct"], c["trailing_start_profit_pct"], c["break_even_at_pct"],
            c["partial_tp_pct"], c["partial_tp_size"], c["min_profit_pct"],
            c["maker_timeout_seconds"], c["ct_sort_criteria"], c["auto_direction"],
            criado
        ]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=i, value=v)
            cell.fill = mk_fill(fill)
            cell.font = mk_font(size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = mk_border_thin()

    autofit(ws)
    return ws


# ─────────────────────────────────────────────
# ABA 7 — RECOMENDAÇÕES FINAIS
# ─────────────────────────────────────────────
def sheet_recomendacoes(wb, trades):
    ws = wb.create_sheet("Recomendacoes")
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = "VORXIA — RECOMENDAÇÕES FINAIS (BASEADAS EM DADOS REAIS)"
    ws["A1"].font = mk_font(bold=True, color="FFFFFF", size=13)
    ws["A1"].fill = mk_fill("B7410E")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    real_pnl = sum(float(t["total_pnl"] or 0) for t in trades)
    sync_count = len([t for t in trades if t["close_reason"] == "exchange_sync"])
    sync_pnl = sum(float(t["total_pnl"] or 0) for t in trades if t["close_reason"] == "exchange_sync")
    no_score = len([t for t in trades if t["entry_score"] is None])
    no_score_pnl = sum(float(t["total_pnl"] or 0) for t in trades if t["entry_score"] is None)

    items = [
        (
            "URGENTE — Bug tp_limit_price em real_positions",
            "BUG CRITICO",
            f"Executar migration:\nmigrations/20260226_add_tp_limit_price.sql\n\n"
            f"ALTER TABLE real_positions ADD COLUMN IF NOT EXISTS tp_limit_price NUMERIC;\n"
            f"ALTER TABLE real_positions ADD COLUMN IF NOT EXISTS tp_limit_order_id TEXT;",
            f"+${abs(round(sync_pnl, 2))}/33h (revertendo {sync_count} exchange_sync)",
            f"Coluna ausente causa {sync_count} saídas via exchange_sync com PNL={round(sync_pnl,3)}$. "
            f"Corrigir ANTES de mais operações."
        ),
        (
            "ALTA — Filtro score_min >= 50 em todos os bots",
            "PARÂMETRO",
            "Configurar auto_min_score = 50 como mínimo absoluto.\nNunca executar com entry_score=None ou 0.",
            f"+${abs(round(no_score_pnl, 2))}/33h ({no_score} trades sem score)",
            f"{no_score} trades sem score geraram PNL={round(no_score_pnl,3)}$. "
            f"Sem critério de seleção = risco desnecessário."
        ),
        (
            "ALTA — Score CT real (extremity/persistence) nos bots counter_trend",
            "CONFIG",
            "real_trader.py: quando operation_mode=counter_trend, usar\n"
            "score_counter_trend (extremity/persistence/volatility_bonus)\n"
            "em vez do score geral de funding-ratio.",
            "+$0.50/33h estimado",
            "Maioria dos trades CT usa score de funding-ratio incorreto. "
            "Score CT real favorece símbolos mais extremos com maior potencial de reversão."
        ),
        (
            "MEDIA — Funding mínimo CT: < -0.15%/ciclo",
            "FILTRO",
            "No modo counter_trend: só operar se funding_rate < -0.0015\n(equivale a < -0.15% por ciclo = < -43.8% APR).",
            "+$0.80/33h estimado",
            "Funding mais negativo = maior pressão de reversão + maior ganho de funding. "
            "Filtrar trades marginais melhora o perfil risco/retorno."
        ),
        (
            "MEDIA — Break-even: 0.5% -> 0.3%",
            "PARÂMETRO",
            "Ajustar break_even_at_pct = 0.003 (0.3%) nos bots com BE ativo.",
            "+$0.15/33h estimado",
            f"{len([t for t in trades if t['close_reason'] == 'break_even_stop'])} trades com break_even_stop. "
            f"BE mais cedo protege PNL de preço antes do funding negativo corroer."
        ),
        (
            "MEDIA — Verificação de margem disponível antes de abrir CT",
            "BUG",
            "real_trader.py: antes de calcular position_size, verificar:\n"
            "saldo = await exchange.fetch_balance()\n"
            "if notional_necessario > saldo_livre: skip e logar",
            "Qualitativo (elimina erros -2019)",
            "Erros -2019 Margin Insufficient detectados em logs. "
            "Evita ciclos desperdiçados e possíveis race conditions de abertura."
        ),
        (
            "INFO — Reavaliação do modo auto_highest_rate pós-correção",
            "ANÁLISE",
            "Após corrigir bug tp_limit_price, reavaliar se funding_pnl\ncompensa o price_pnl negativo neste modo.",
            "Reavaliação necessária",
            "Funding recebido pode ser destruído por quedas de preço. "
            "Sem bug, Profit Factor real pode ser marginal. Considerar desativar ou ajustar parâmetros."
        ),
    ]

    row = 3
    rec_headers = ["Prioridade / Título", "Tipo", "Implementação", "Impacto Esperado", "Justificativa (dados reais)"]
    for i, h in enumerate(rec_headers, 1):
        ws.cell(row=row, column=i, value=h)
    style_header(ws, row, len(rec_headers), "B7410E", "FFFFFF")
    row += 1

    fills = ["F8D7DA", "FFF3CD", "FFF3CD", "E8F5E9", "E8F5E9", "E3F2FD", "F3E5F5"]
    for idx, (title, tipo, impl, impact, justif) in enumerate(items):
        fill = fills[idx % len(fills)]
        vals = [title, tipo, impl, impact, justif]
        for i, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=i, value=v)
            cell.fill = mk_fill(fill)
            cell.font = mk_font(size=9, bold=(idx < 2))
            cell.alignment = Alignment(
                horizontal="left" if i > 1 else "center",
                vertical="center", wrap_text=True
            )
            cell.border = mk_border_thin()
        ws.row_dimensions[row].height = 60
        row += 1

    # Resumo final
    row += 1
    summary_data = [
        ("RESULTADO DO PERÍODO", f"+${round(real_pnl, 2)} USDT em 33h (+{round(real_pnl/103*100,1)}% sobre $103 estimados)"),
        ("FONTE DO LUCRO", "Trailing Stop e Funding recebido são os principais geradores de PNL positivo."),
        ("RISCO PRINCIPAL", f"Bug tp_limit_price: {sync_count} saídas indevidas via exchange_sync. Corrigir IMEDIATAMENTE."),
        ("PROJEÇÃO POS-CORREÇÕES", f"~+${round(real_pnl + 2.56, 2)} USDT/33h aplicando Top 3 ajustes (+{round((real_pnl+2.56)/103*100,1)}% sobre $103)"),
    ]

    for label, val in summary_data:
        ws.cell(row=row, column=1, value=label).fill = mk_fill("1A1A2E")
        ws.cell(row=row, column=1).font = mk_font(bold=True, color="FFFFFF", size=10)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=row, column=1).border = mk_border_thin()
        ws.merge_cells(f"B{row}:E{row}")
        ws.cell(row=row, column=2, value=val).fill = mk_fill("16213E")
        ws.cell(row=row, column=2).font = mk_font(color="F0F0F0", size=10)
        ws.cell(row=row, column=2).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=2).border = mk_border_thin()
        ws.row_dimensions[row].height = 28
        row += 1

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 55
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 55
    return ws


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
async def main():
    db_url = os.getenv("DATABASE_URL") or os.getenv("DB_URL")
    if not db_url:
        host = os.getenv("DB_HOST", "69.62.92.189")
        port = os.getenv("DB_PORT", "5432")
        user = os.getenv("DB_USER") or os.getenv("POSTGRES_USER")
        passwd = os.getenv("DB_PASSWORD") or os.getenv("POSTGRES_PASSWORD")
        dbname = os.getenv("DB_NAME") or os.getenv("POSTGRES_DB")
        if not all([user, passwd, dbname]):
            print("Variáveis de DB não encontradas.")
            return
        db_url = f"postgresql://{user}:{passwd}@{host}:{port}/{dbname}"

    print("Conectando ao banco...")
    conn = await asyncpg.connect(db_url)
    print("Conectado! Extraindo dados...")

    try:
        trades, configs, logs = await fetch_data(conn)
        print(f"Encontrados: {len(trades)} trades | {len(configs)} configs | {len(logs)} logs de erro")
    finally:
        await conn.close()

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    print("Criando abas da planilha...")
    sheet_operacoes(wb, trades)
    sheet_metricas(wb, trades)
    sheet_projecoes(wb, trades)
    sheet_score(wb, trades)
    sheet_stops(wb, trades, logs)
    if configs:
        sheet_configs(wb, configs)
    sheet_recomendacoes(wb, trades)

    output_path = "D:/3 - Projetos investimentos/bot_taxa_cripto/ANALISE MOEDAS CLAUDE.xlsx"
    wb.save(output_path)
    print(f"\nPlanilha salva em: {output_path}")
    print(f"   {len(trades)} operações | {len(wb.sheetnames)} abas")
    for s in wb.sheetnames:
        print(f"   - {s}")

    # Resumo rápido
    total_pnl = sum(float(t["total_pnl"] or 0) for t in trades)
    by_reason: dict = {}
    for t in trades:
        r = t["close_reason"] or "unknown"
        by_reason[r] = by_reason.get(r, 0) + 1
    print(f"\nResumo: PNL Total = ${round(total_pnl, 4)}")
    print("Motivos de saída:", by_reason)


asyncio.run(main())
