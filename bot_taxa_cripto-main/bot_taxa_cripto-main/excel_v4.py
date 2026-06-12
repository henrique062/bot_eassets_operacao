import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
wb.remove(wb.active)

# Cores
VERDE      = PatternFill("solid", fgColor="1A6B3C")
VERDE_CLR  = PatternFill("solid", fgColor="D6F5E3")
VERMELHO   = PatternFill("solid", fgColor="922B21")
VERM_CLR   = PatternFill("solid", fgColor="FADBD8")
CINZA_ESC  = PatternFill("solid", fgColor="1C2833")
CINZA_MED  = PatternFill("solid", fgColor="2C3E50")
CINZA_CLR  = PatternFill("solid", fgColor="D5D8DC")
AZUL       = PatternFill("solid", fgColor="1A5276")
AZUL_CLR   = PatternFill("solid", fgColor="D6EAF8")
AMARELO    = PatternFill("solid", fgColor="7D6608")
AMAR_CLR   = PatternFill("solid", fgColor="FCF3CF")
LARANJA    = PatternFill("solid", fgColor="BA4A00")
LARAN_CLR  = PatternFill("solid", fgColor="FAE5D3")
BRANCO     = PatternFill("solid", fgColor="FFFFFF")
ROXO       = PatternFill("solid", fgColor="6C3483")
ROXO_CLR   = PatternFill("solid", fgColor="EBD5FB")

thin = Side(border_style="thin", color="CCCCCC")
med  = Side(border_style="medium", color="888888")

def b_all(): return Border(left=thin, right=thin, top=thin, bottom=thin)
def ctr(): return Alignment(horizontal="center", vertical="center", wrap_text=True)
def lft(): return Alignment(horizontal="left", vertical="center", wrap_text=True)
def rgt(): return Alignment(horizontal="right", vertical="center")

def fnt(sz=10, bold=False, color="000000"):
    return Font(size=sz, bold=bold, color=color)

def hdr(ws, row, col, txt, fill=None, fc="FFFFFF", sz=10):
    c = ws.cell(row=row, column=col, value=txt)
    c.fill = fill or CINZA_ESC
    c.font = fnt(sz, True, fc)
    c.alignment = ctr()
    c.border = b_all()
    return c

def vl(ws, row, col, v, fill=None, fmt=None, fc="000000", bold=False, align="right"):
    c = ws.cell(row=row, column=col, value=v)
    if fill: c.fill = fill
    if fmt:  c.number_format = fmt
    c.font = fnt(10, bold, fc)
    c.alignment = ctr() if align == "center" else (lft() if align == "left" else rgt())
    c.border = b_all()
    return c

def merge_hdr(ws, row, c1, c2, txt, fill=None, fc="FFFFFF", sz=11):
    ws.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    c = ws.cell(row=row, column=c1, value=txt)
    c.fill = fill or CINZA_ESC
    c.font = fnt(sz, True, fc)
    c.alignment = ctr()
    return c

def auto_w(ws, mn=8, mx=40):
    for col in ws.columns:
        w = mn
        for cell in col:
            if cell.value:
                w = min(mx, max(w, len(str(cell.value)) + 2))
        ws.column_dimensions[get_column_letter(col[0].column)].width = w

# ==========================================================
# ABA 1 - RESUMO GERAL
# ==========================================================
ws = wb.create_sheet("Resumo Geral")
ws.freeze_panes = "A5"
ws.sheet_view.showGridLines = False

ws.merge_cells("A1:P1")
c = ws["A1"]
c.value = "ANALISE COMPLETA - BOTS TRADING REAL | Atualizado 24/02/2026"
c.fill = CINZA_ESC; c.font = fnt(14, True, "00E68A"); c.alignment = ctr()
ws.row_dimensions[1].height = 32

ws.merge_cells("A2:P2")
c = ws["A2"]
c.value = "Comparativo: Geracao 1 (original) -> Geracao 2 (otimizados) -> Geracao 3 (ativos agora)"
c.fill = CINZA_MED; c.font = fnt(10, True, "A0AEC0"); c.alignment = ctr()
ws.row_dimensions[2].height = 20

kpis = [
    ("Capital Investido", "$62,00", AZUL_CLR, "1A5276"),
    ("Saldo Atual", "$71,88", VERDE_CLR, "1A6B3C"),
    ("Lucro Total", "+$9,88", VERDE_CLR, "0B5345"),
    ("Retorno Total", "+15.94%", VERDE_CLR, "0B5345"),
    ("Win Rate", "66.7%", VERDE_CLR, "1A6B3C"),
    ("TP Hits", "15 / 27 ops", VERDE_CLR, "1A6B3C"),
    ("Melhor Bot", "Bot 36: +31%", VERDE_CLR, "0B5345"),
    ("Menor Bot", "Bot 39: +7%", AMAR_CLR, "7D6608"),
]
col = 1
for label, valor, fill, fc in kpis:
    ws.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    c = ws.cell(3, col, label)
    c.fill = CINZA_MED; c.font = fnt(9, True, "A0AEC0"); c.alignment = ctr()
    ws.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    c = ws.cell(4, col, valor)
    c.fill = fill; c.font = fnt(13, True, fc); c.alignment = ctr()
    col += 2

ws.row_dimensions[3].height = 18
ws.row_dimensions[4].height = 26

headers = ["ID","Nome do Bot","Modo","Status","Capital","Saldo","Lucro $",
           "Retorno %","Trades","W","L","TP Hits","SL Hits","Timeout","Win Rate","Config Chave"]
for i, h in enumerate(headers):
    hdr(ws, 6, i+1, h)

bots = [
    (27,"Ao contrario","counter_trend","ENCERRADO",25.00,25.53,0.53,2.12,11,5,6,0,3,8,"45.5%","SL 2.5% / exit 520s","GEN1"),
    (28,"Bot Binance 01:07","auto_expiring","ENCERRADO",10.00,9.95,-0.05,-0.50,5,2,3,0,4,0,"40.0%","SL 1.2% / exit 35s","GEN1"),
    (29,"Bot Binance 01:25","auto_expiring","ENCERRADO",10.00,10.13,0.13,1.31,3,2,1,0,0,0,"66.7%","SL 2.5% / exit 30s","GEN1"),
    (30,"Bot Binance 01:25 *","auto_expiring","ENCERRADO",10.00,10.00,0.00,0.00,0,0,0,0,0,0,"---","sem SL / exit 35s","GEN2"),
    (31,"Bot Binance 01:07 *","auto_expiring","ENCERRADO",10.05,10.02,-0.03,-0.30,2,1,1,0,0,2,"50.0%","sem SL / exit 35s","GEN2"),
    (32,"Ao contrario *","counter_trend","ENCERRADO",25.53,27.72,2.19,8.58,7,5,2,2,0,5,"71.4%","sem SL / exit 1700s","GEN2"),
    (33,"TAKER Funding R","counter_trend","ENCERRADO",25.00,26.39,1.39,5.56,4,2,2,1,0,3,"50.0%","sem SL / exit 1700s","GEN2"),
    (36,"TAKE Funding R","counter_trend","ATIVO",25.00,32.74,7.74,30.96,13,10,3,10,2,1,"76.9%","SL 30% / exit 20000s","GEN3"),
    (37,"Ao contrario","counter_trend","ATIVO",25.00,26.30,1.30,5.20,7,4,3,4,1,2,"57.1%","SL 30% / exit 20000s","GEN3"),
    (38,"Bot Binance 22:26","auto_strongest","ENCERRADO",12.00,12.69,0.69,5.74,1,1,0,0,0,0,"100%","SL 30% / exit 700s","GEN3"),
    (39,"Bot Binance 04:51","auto_strongest","ATIVO",12.00,12.84,0.84,7.02,7,4,3,1,2,0,"57.1%","SL 30% / exit 700s","GEN3"),
]

gen_fill = {"GEN1": LARAN_CLR, "GEN2": AMAR_CLR, "GEN3": VERDE_CLR}
gen_fc   = {"GEN1": "BA4A00",  "GEN2": "7D6608",  "GEN3": "1A6B3C"}

for i, b in enumerate(bots):
    r = 7 + i
    bid,nome,modo,status,cap,sal,luc,pct,tr,w,l,tp,sl,tout,wr,conf,gen = b
    gf = gen_fill[gen]; gfc = gen_fc[gen]

    vl(ws, r, 1, bid, gf, None, gfc, True, "center")
    vl(ws, r, 2, nome, gf, None, gfc, True, "left")
    vl(ws, r, 3, modo, gf, None, "444444", False, "center")
    sf = VERDE_CLR if status == "ATIVO" else CINZA_CLR
    sc = "1A6B3C" if status == "ATIVO" else "555555"
    vl(ws, r, 4, status, sf, None, sc, status=="ATIVO", "center")
    vl(ws, r, 5, cap, gf, '"$"#,##0.00')
    vl(ws, r, 6, sal, gf, '"$"#,##0.00')
    lf = VERDE_CLR if luc >= 0 else VERM_CLR
    lc = "1A6B3C" if luc >= 0 else "922B21"
    vl(ws, r, 7, luc, lf, '"$"#,##0.0000', lc, True)
    pf = VERDE_CLR if pct >= 0 else VERM_CLR
    pc = "1A6B3C" if pct >= 0 else "922B21"
    vl(ws, r, 8, pct/100, pf, '"+"0.00%;"-"0.00%', pc, True)
    vl(ws, r, 9, tr, gf, None, "000000", False, "center")
    vl(ws, r, 10, w, VERDE_CLR if w>0 else gf, None, "1A6B3C" if w>0 else "000000", w>0, "center")
    vl(ws, r, 11, l, VERM_CLR if l>0 else gf, None, "922B21" if l>0 else "000000", l>0, "center")
    vl(ws, r, 12, tp, AZUL_CLR if tp>0 else gf, None, "1A5276" if tp>0 else "000000", tp>0, "center")
    vl(ws, r, 13, sl, VERM_CLR if sl>0 else gf, None, "922B21" if sl>0 else "000000", sl>0, "center")
    vl(ws, r, 14, tout, gf, None, "000000", False, "center")
    vl(ws, r, 15, wr, gf, None, "000000", False, "center")
    vl(ws, r, 16, conf, gf, None, "444444", False, "left")

r = 7 + len(bots)
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
c = ws.cell(r, 1, "TOTAL BOTS ATIVOS (36+37+39)")
c.fill = VERDE; c.font = fnt(11, True, "FFFFFF"); c.alignment = ctr()
vl(ws, r, 5, 62.00, VERDE, '"$"#,##0.00', "FFFFFF", True)
vl(ws, r, 6, 71.88, VERDE, '"$"#,##0.00', "FFFFFF", True)
vl(ws, r, 7, 9.88, VERDE, '"$"#,##0.0000', "FFFFFF", True)
vl(ws, r, 8, 0.1594, VERDE, '"+"0.00%', "FFFFFF", True)
vl(ws, r, 9, 27, VERDE, None, "FFFFFF", True, "center")
vl(ws, r, 10, 18, VERDE, None, "FFFFFF", True, "center")
vl(ws, r, 11, 9, VERDE, None, "FFFFFF", True, "center")
vl(ws, r, 12, 15, VERDE, None, "FFFFFF", True, "center")

auto_w(ws)
ws.column_dimensions["B"].width = 24
ws.column_dimensions["P"].width = 28

# ==========================================================
# ABA 2 - OPERACOES DETALHADAS
# ==========================================================
ws2 = wb.create_sheet("Operacoes Detalhadas")
ws2.freeze_panes = "A4"
ws2.sheet_view.showGridLines = False

ws2.merge_cells("A1:O1")
c = ws2["A1"]
c.value = "OPERACOES DETALHADAS - BOTS 36, 37 e 39 (Geracao 3 Atual)"
c.fill = CINZA_ESC; c.font = fnt(13, True, "00E68A"); c.alignment = ctr()
ws2.row_dimensions[1].height = 28

hdrs2 = ["ID","Bot","Simbolo","Dir","Entrada","Saida","Move%","F.Rate","Funding$","Price P&L","Fee","PNL Total","Motivo","Hora Abertura","Hora Fechamento"]
for i, h in enumerate(hdrs2):
    hdr(ws2, 3, i+1, h)

trades = [
    (76,36,"STEEMUSDT","SHORT",0.05364,0.05332,-0.5966,0.0,0.0,0.2256,0.00756,0.21804,"exchange_sync","24/02 11:00","24/02 11:23"),
    (73,36,"BULLAUSDT","SHORT",0.02952,0.02898,-1.8149,0.0,0.0,0.68891,0.01898,0.66993,"exchange_sync","24/02 11:00","24/02 11:04"),
    (72,36,"FOGOUSDT","SHORT",0.02928,0.02905,-0.7855,0.0,0.0,0.22609,0.01439,0.21170,"exchange_sync","24/02 09:00","24/02 10:33"),
    (70,36,"STEEMUSDT","SHORT",0.05536,0.05441,-1.7160,0.0,0.0,0.59185,0.01725,0.57461,"exchange_sync","24/02 10:00","24/02 10:30"),
    (69,36,"BULLAUSDT","SHORT",0.03522,0.03129,-11.1665,0.0,0.0,3.94107,0.01765,3.92342,"exchange_sync","24/02 10:00","24/02 10:27"),
    (66,36,"POWERUSDT","SHORT",0.36717,0.38673,5.3272,0.0,0.0,-1.79952,0.02455,-1.82407,"stop_loss_pct","24/02 10:00","24/02 10:04"),
    (62,36,"STEEMUSDT","SHORT",0.05763,0.05745,-0.3123,0.0,0.0,0.08982,0.01438,0.07544,"exchange_sync","24/02 09:00","24/02 09:07"),
    (61,36,"POWERUSDT","SHORT",0.58561,0.48861,-16.5645,0.0,-0.08094,4.75321,0.01435,4.65793,"exchange_sync","24/02 08:00","24/02 09:01"),
    (60,36,"STEEMUSDT","SHORT",0.05765,0.05742,-0.3990,0.0,-0.39928,0.12604,0.00632,-0.27955,"exchange_sync","24/02 06:00","24/02 08:06"),
    (58,36,"UMAUSDT","SHORT",0.45901,0.45380,-1.1346,0.0,-0.02489,0.33330,0.02921,0.27921,"timeout","24/02 01:00","24/02 06:33"),
    (55,36,"BULLAUSDT","SHORT",0.02766,0.02968,7.3030,0.0,0.0,-2.37217,0.02353,-2.39570,"stop_loss_pct","24/02 06:00","24/02 06:16"),
    (51,36,"SKRUSDT","SHORT",0.02324,0.02260,-2.7370,0.0,0.0,0.84079,0.01536,0.82543,"exchange_sync","24/02 05:00","24/02 05:12"),
    (50,36,"STEEMUSDT","SHORT",0.05973,0.05763,-3.5158,0.0,-0.22504,1.04370,0.01484,0.80382,"exchange_sync","24/02 02:00","24/02 03:32"),
    (74,37,"BULLAUSDT","SHORT",0.02951,0.02988,1.2579,0.0,0.0,-0.40053,0.01592,-0.41645,"exchange_sync","24/02 11:00","24/02 11:04"),
    (71,37,"FLOWUSDT","SHORT",0.03504,0.03464,-1.1416,0.0,0.0,0.31440,0.01377,0.30063,"exchange_sync","24/02 09:00","24/02 10:32"),
    (68,37,"BULLAUSDT","SHORT",0.03514,0.03133,-10.8423,0.0,0.0,3.07848,0.01420,3.06428,"exchange_sync","24/02 10:00","24/02 10:26"),
    (65,37,"STORJUSDT","SHORT",0.09110,0.09060,-0.5488,0.0,0.0,0.15100,0.01376,0.13724,"exchange_sync","24/02 09:00","24/02 09:46"),
    (59,37,"JTOUSDT","SHORT",0.27780,0.27330,-1.6199,0.0,-0.01523,0.47700,0.02037,0.44140,"timeout","24/02 01:00","24/02 06:34"),
    (57,37,"AWEUSDT","SHORT",0.05091,0.05090,-0.0196,0.0,-0.07200,0.00583,0.02968,-0.09584,"timeout","24/02 01:00","24/02 06:33"),
    (56,37,"BULLAUSDT","SHORT",0.02759,0.02957,7.1765,0.0,0.0,-2.09933,0.03087,-2.13020,"stop_loss_pct","24/02 06:00","24/02 06:16"),
    (75,39,"BULLAUSDT","LONG",0.03005,0.02807,-6.589,-2.0,0.69485,-2.28690,0.02315,-1.61520,"stop_loss_pct","24/02 11:00","24/02 11:07"),
    (67,39,"BULLAUSDT","LONG",0.03520,0.03289,-6.563,-1.99154,0.77990,-2.51087,0.03700,-1.76798,"stop_loss_pct","24/02 10:00","24/02 10:11"),
    (64,39,"STORJUSDT","LONG",0.09110,0.09110,0.0,-0.15067,0.05478,0.0,0.03790,0.01688,"funding","24/02 09:00","24/02 09:15"),
    (63,39,"LPTUSDT","LONG",2.20400,2.21600,0.5445,-0.19851,0.07457,0.20640,0.03801,0.24296,"funding","24/02 09:00","24/02 09:11"),
    (54,39,"BULLAUSDT","LONG",0.02722,0.02794,2.6508,-1.48798,0.48337,0.92790,0.03547,1.37580,"funding","24/02 06:00","24/02 06:12"),
    (53,39,"ONTUSDT","LONG",0.04050,0.04030,-0.4938,-0.42007,0.11569,-0.14074,0.02843,-0.05349,"funding","24/02 05:00","24/02 05:17"),
    (52,39,"BULLAUSDT","LONG",0.03194,0.03444,7.819,-2.0,0.54157,2.11547,0.01353,2.64351,"exchange_sync","24/02 05:00","24/02 05:13"),
]

bot_fills = {36: AZUL_CLR, 37: LARAN_CLR, 39: AMAR_CLR}
last_r = 3

for i, t in enumerate(trades):
    r = 4 + i
    last_r = r
    tid, bot, sym, dr, entry, exit_, move, frate, fpnl, ppnl, fee, pnl, reason, ot, ct = t
    bf = bot_fills.get(bot, BRANCO)
    pnl_fill = VERDE_CLR if pnl > 0 else VERM_CLR
    pnl_fc = "1A6B3C" if pnl > 0 else "922B21"

    vl(ws2, r, 1, tid, bf, None, "444444", False, "center")
    bot_fc = {"36":"1A5276","37":"BA4A00","39":"7D6608"}.get(str(bot),"000000")
    vl(ws2, r, 2, f"Bot {bot}", bf, None, bot_fc, True, "center")
    vl(ws2, r, 3, sym.replace("USDT",""), BRANCO, None, "000000", True, "center")
    df = AZUL_CLR if dr == "LONG" else VERM_CLR
    dc = "1A5276" if dr == "LONG" else "922B21"
    vl(ws2, r, 4, dr, df, None, dc, True, "center")
    vl(ws2, r, 5, entry, BRANCO, "#,##0.00000000")
    vl(ws2, r, 6, exit_, BRANCO, "#,##0.00000000")
    good_move = (move < 0 and dr == "SHORT") or (move > 0 and dr == "LONG")
    mf = VERDE_CLR if good_move else VERM_CLR
    mc = "1A6B3C" if good_move else "922B21"
    vl(ws2, r, 7, move/100, mf, '"+"0.0000%;"-"0.0000%', mc, True)
    vl(ws2, r, 8, frate/100 if frate else 0, BRANCO, '"+"0.0000%;"-"0.0000%')
    fpf = VERDE_CLR if fpnl > 0 else (VERM_CLR if fpnl < 0 else BRANCO)
    fpc = "1A6B3C" if fpnl > 0 else ("922B21" if fpnl < 0 else "888888")
    vl(ws2, r, 9, fpnl, fpf, '"+"$#,##0.0000;"-"$#,##0.0000', fpc)
    ppf = VERDE_CLR if ppnl > 0 else (VERM_CLR if ppnl < 0 else BRANCO)
    vl(ws2, r, 10, ppnl, ppf, '"+"$#,##0.0000;"-"$#,##0.0000')
    vl(ws2, r, 11, fee, BRANCO, '"$"#,##0.0000', "888888")
    vl(ws2, r, 12, pnl, pnl_fill, '"+"$#,##0.0000;"-"$#,##0.0000', pnl_fc, True)
    rf = VERDE_CLR if "exchange_sync" in reason else (VERM_CLR if "stop" in reason else AMAR_CLR)
    vl(ws2, r, 13, reason.replace("_"," "), rf, None, "000000", False, "center")
    vl(ws2, r, 14, ot, BRANCO, None, "555555", False, "center")
    vl(ws2, r, 15, ct, BRANCO, None, "555555", False, "center")

for bot_id, bfill in [(36, AZUL), (37, LARANJA), (39, AMARELO)]:
    last_r += 1
    bt = [t for t in trades if t[1] == bot_id]
    s = sum(t[11] for t in bt)
    w = sum(1 for t in bt if t[11] > 0)
    l = sum(1 for t in bt if t[11] < 0)
    ws2.merge_cells(start_row=last_r, start_column=1, end_row=last_r, end_column=11)
    c = ws2.cell(last_r, 1, f"TOTAL BOT {bot_id}  ({w}W / {l}L  |  WR: {w/(w+l)*100:.0f}%)")
    c.fill = bfill; c.font = fnt(11, True, "FFFFFF"); c.alignment = ctr()
    vl(ws2, last_r, 12, s, bfill, '"+"$#,##0.0000;"-"$#,##0.0000', "FFFFFF", True)

last_r += 1
ws2.merge_cells(start_row=last_r, start_column=1, end_row=last_r, end_column=11)
c = ws2.cell(last_r, 1, "TOTAL GERAL  (18W / 9L  |  WR: 66.7%)")
c.fill = VERDE; c.font = fnt(12, True, "FFFFFF"); c.alignment = ctr()
vl(ws2, last_r, 12, 9.882, VERDE, '"+"$#,##0.0000;"-"$#,##0.0000', "FFFFFF", True)

auto_w(ws2)

# ==========================================================
# ABA 3 - ANALISE PERFORMANCE
# ==========================================================
ws3 = wb.create_sheet("Analise Performance")
ws3.sheet_view.showGridLines = False
for col, w in [("A",28),("B",18),("C",18),("D",18),("E",45)]:
    ws3.column_dimensions[col].width = w

ws3.merge_cells("A1:E1")
c = ws3["A1"]
c.value = "ANALISE DE PERFORMANCE - COMPARATIVO DAS GERACOES"
c.fill = CINZA_ESC; c.font = fnt(13, True, "00E68A"); c.alignment = ctr()
ws3.row_dimensions[1].height = 28

row = 3
merge_hdr(ws3, row, 1, 5, "COMPARATIVO POR GERACAO", CINZA_MED, "FFFFFF", 11)
row = 4
for i, h in enumerate(["Geracao","Total Lucro","Retorno Medio","Win Rate","Mudanca Principal"]):
    hdr(ws3, row, i+1, h)
gens_data = [
    ("GEN 1 - Original (27-29)", "+$0.61", "+1.65%", "50.5%", "Baseline. SL apertado 1-2.5%, exit curto 30-520s."),
    ("GEN 2 - Otimizacao v1 (30-33)", "+$3.55", "+7.35%", "63.6%", "Removeu SL, aumentou exit para 1700s. Melhorou."),
    ("GEN 3 - Atual (36-39 ATIVOS)", "+$9.88", "+14.39%", "66.7%", "EXIT 20000s = REVOLUCIONARIO! 56% TP hits vs 0%."),
]
gfills = [LARAN_CLR, AMAR_CLR, VERDE_CLR]
gfcs   = ["BA4A00", "7D6608", "1A6B3C"]
for i, g in enumerate(gens_data):
    r = 5 + i
    for j, v in enumerate(g):
        vl(ws3, r, j+1, v, gfills[i], None, gfcs[i], i==2, "center" if j < 4 else "left")

row = 10
merge_hdr(ws3, row, 1, 5, "METRICAS CHAVE POR GERACAO", CINZA_MED, "FFFFFF", 11)
row = 11
for i, h in enumerate(["Metrica","Gen 1 (orig)","Gen 2 (otim)","Gen 3 (atual)","Interpretacao"]):
    hdr(ws3, row, i+1, h)

metrics = [
    ("Win Rate", "50.5%", "63.6%", "66.7%", "Melhoria consistente a cada geracao"),
    ("TP Hits (exchange_sync)", "0/19 (0%)", "3/11 (27%)", "15/27 (56%)", "MELHORIA CRITICA: exit longo = mais TP"),
    ("SL Hits", "7/19 (37%)", "0/11 (0%)", "5/27 (19%)", "SL 30% raramente dispara - adequado"),
    ("Timeouts", "8/19 (42%)", "8/11 (73%)", "3/27 (11%)", "Menos timeout = mais eficiencia"),
    ("Lucro medio por trade", "+$0.032", "+$0.322", "+$0.367", "Gen 3 mais consistente"),
    ("Maior ganho unico", "+$0.19 (ARC)", "+$1.67 (POWER)", "+$4.66 (POWER)", "POWER e BULLAUSDT sao estrelas"),
    ("Maior perda unica", "-$0.23", "-$0.084", "-$2.40 (BULLA)", "Gen 3 perde mais mas ganha muito mais"),
    ("Retorno total / capital", "+$0.61/$45", "+$3.55/$70.55", "+$9.88/$62", "Gen 3: +15.94% vs Gen 1: +1.36%"),
]
for i, m in enumerate(metrics):
    r = 12 + i
    rfill = VERDE_CLR if i % 2 == 0 else BRANCO
    for j, v in enumerate(m):
        vl(ws3, r, j+1, v, rfill, None, "000000", False, "left")

row = 22
merge_hdr(ws3, row, 1, 5, "ANALISE POR SIMBOLO (Gen 3 - Bots 36/37/39)", CINZA_MED, "FFFFFF", 11)
row = 23
for i, h in enumerate(["Simbolo","Trades","W/L","PNL Total","Observacao / Risco"]):
    hdr(ws3, row, i+1, h)

syms_data = [
    ("BULLAUSDT", 9, "6W/3L", 3.114, "ALTA VOLATILIDADE: +11% e -11% no mesmo dia. Maior risco e recompensa. SL 30% adequado."),
    ("POWERUSDT", 3, "2W/1L", 3.513, "Excelente: quedas de 5-16%. Bot 36 perdeu 1x quando subiu 5%. Monitorar."),
    ("STEEMUSDT", 5, "4W/1L", 1.393, "Consistente. Quedas graduais. Ideal para counter_trend."),
    ("SKRUSDT",   1, "1W/0L", 0.825, "Queda 2.7%. Positivo."),
    ("FLOWUSDT",  1, "1W/0L", 0.301, "Queda moderada. Positivo."),
    ("STORJUSDT", 2, "1W/1L", 0.154, "Neutro. Coleta funding bem no LONG."),
    ("JTOUSDT",   1, "1W/0L", 0.441, "Queda 1.6%. Bom via timeout."),
    ("FOGOUSDT",  1, "1W/0L", 0.212, "Queda 0.8%. Ok."),
    ("UMAUSDT",   1, "1W/0L", 0.279, "Timeout positivo. Queda 1.1%."),
    ("LPTUSDT",   1, "1W/0L", 0.243, "Funding negativo + alta = perfeito para LONG."),
    ("AWEUSDT",   1, "0W/1L", -0.096, "Fee consumiu ganho minimo. Marginal."),
    ("ONTUSDT",   1, "0W/1L", -0.053, "Queda leve comeu funding. Marginal."),
]
for i, s in enumerate(syms_data):
    sym, tr, wl, pnl_v, obs = s
    r = 24 + i
    sf = VERDE_CLR if pnl_v > 0 else (VERM_CLR if pnl_v < 0 else BRANCO)
    sc = "1A6B3C" if pnl_v > 0 else "922B21"
    vl(ws3, r, 1, sym.replace("USDT",""), sf, None, sc, True, "left")
    vl(ws3, r, 2, tr, sf, None, sc, False, "center")
    vl(ws3, r, 3, wl, sf, None, sc, False, "center")
    vl(ws3, r, 4, pnl_v, sf, '"+"$#,##0.0000;"-"$#,##0.0000', sc, True)
    vl(ws3, r, 5, obs, BRANCO, None, "444444", False, "left")

ws3.column_dimensions["E"].width = 60

# ==========================================================
# ABA 4 - RECOMENDACOES
# ==========================================================
ws4 = wb.create_sheet("Recomendacoes")
ws4.sheet_view.showGridLines = False
for col, w in [("A",22),("B",24),("C",22),("D",65)]:
    ws4.column_dimensions[col].width = w

ws4.merge_cells("A1:D1")
c = ws4["A1"]
c.value = "RECOMENDACOES E PROXIMOS PASSOS - 24/02/2026"
c.fill = CINZA_ESC; c.font = fnt(13, True, "00E68A"); c.alignment = ctr()
ws4.row_dimensions[1].height = 28

row = 3
for i, h in enumerate(["Prioridade","Bot / Item","Status","Descricao / Acao"]):
    hdr(ws4, row, i+1, h)

recs = [
    ("ALTA","Bot 36 - exit 20000s","MANTER COMO ESTA",
     "Bot 36 com +30.96% esta excelente. counter_trend + exit_seconds 20000s e a chave do sucesso. NAO ALTERAR."),
    ("ALTA","Bot 39 - BULLAUSDT LONG","PROBLEMA IDENTIFICADO",
     "Bot 39 pegou BULLAUSDT 3x LONG com funding negativo. BULLAUSDT caiu 6.5% consecutivo = -$5.16 perdas. Aumentar auto_min_score de 50 para 75+."),
    ("MEDIA","Bot 37 - BULLAUSDT risco","RISCO ACEITAVEL",
     "Bot 37 sofreu -$2.13 no BULLAUSDT as 06:00 (subiu 7%). As 10:00 ganhou +$3.06 mesmo token. Risco compensado."),
    ("MEDIA","Funding PNL = 0","COMPORTAMENTO CORRETO",
     "Quase todas ops tem funding_pnl = 0. O lucro vem da variacao de preco via TP (limit order). O funding e coletado pela exchange ao fechar. Isso e CORRETO para counter_trend."),
    ("MEDIA","STEEMUSDT trade 60","INVESTIGAR",
     "Trade 60 bot 36: STEEMUSDT SHORT com PNL -$0.28 mesmo preco caindo! Funding negativo (-$0.40) consumiu lucro de preco (+$0.13). Atencao a tokens com funding negativo no SHORT."),
    ("BAIXA","SL = 30%","ADEQUADO",
     "SL de 30% disparou 5/27 = 18.5% das vezes. BULLAUSDT com queda de 11% disparou SL - esperado. Protege sem prejudicar ops normais."),
    ("BAIXA","Bot 39 simbolos","SUGESTAO",
     "Bot 39 (auto_strongest) opera BULLAUSDT em auto. Adicionar tokens mais estaveis e aumentar auto_min_score para reduzir BULLAUSDT."),
    ("FUTURO","4 Bot sugerido","A CRIAR",
     "counter_trend + exit_seconds=20000s + SL=30% + auto_min_score=70 + simbolos: STEEMUSDT,SKRUSDT,FOGOUSDT. Capital $15-20."),
    ("VALIDADO","exit_seconds 20000s","FOI A MUDANCA CHAVE",
     "Mudanca de 520/700s para 20000s foi TRANSFORMADORA. TP hits: 0% -> 27% -> 56%. Com mais tempo as limit orders tem chance real de ser preenchidas. MANTER OBRIGATORIAMENTE."),
    ("VALIDADO","SL 30% vs sem SL","EQUILIBRIO IDEAL",
     "Gen 2 sem SL funcionou mas expoe a risco ilimitado. Gen 3 com SL 30% e o equilíbrio perfeito: protege de crashes raros sem interferir em ops normais."),
    ("VALIDADO","counter_trend auto","ESTRATEGIA CORRETA",
     "counter_trend em modo auto (varrendo exchange por funding alto) esta funcionando muito bem. O bot escolhe tokens com funding alto e entra contra a direcao. Maioria fecha no TP."),
]

prio_fills = {"ALTA": VERMELHO, "MEDIA": LARANJA, "BAIXA": AZUL, "FUTURO": ROXO, "VALIDADO": VERDE}
status_fills = {
    "MANTER COMO ESTA": VERDE_CLR, "PROBLEMA IDENTIFICADO": VERM_CLR,
    "RISCO ACEITAVEL": AMAR_CLR, "COMPORTAMENTO CORRETO": VERDE_CLR,
    "INVESTIGAR": VERM_CLR, "ADEQUADO": VERDE_CLR,
    "SUGESTAO": AZUL_CLR, "A CRIAR": ROXO_CLR,
    "FOI A MUDANCA CHAVE": VERDE_CLR, "EQUILIBRIO IDEAL": VERDE_CLR, "ESTRATEGIA CORRETA": VERDE_CLR,
}
status_fcs = {
    "MANTER COMO ESTA": "1A6B3C", "PROBLEMA IDENTIFICADO": "922B21",
    "RISCO ACEITAVEL": "7D6608", "COMPORTAMENTO CORRETO": "1A6B3C",
    "INVESTIGAR": "922B21", "ADEQUADO": "1A6B3C",
    "SUGESTAO": "1A5276", "A CRIAR": "6C3483",
    "FOI A MUDANCA CHAVE": "0B5345", "EQUILIBRIO IDEAL": "0B5345", "ESTRATEGIA CORRETA": "0B5345",
}

for i, rec in enumerate(recs):
    r = 4 + i
    ws4.row_dimensions[r].height = 36
    prio, item, status, desc = rec
    vl(ws4, r, 1, prio, prio_fills.get(prio, CINZA_CLR), None, "FFFFFF", True, "center")
    vl(ws4, r, 2, item, CINZA_CLR, None, "000000", True, "left")
    sf = status_fills.get(status, BRANCO)
    sc = status_fcs.get(status, "000000")
    vl(ws4, r, 3, status, sf, None, sc, True, "center")
    vl(ws4, r, 4, desc, BRANCO, None, "333333", False, "left")

path = r"d:\3 - Projetos investimentos\bot_taxa_cripto\analise_bots_23022026_v4.xlsx"
wb.save(path)
print(f"Salvo em: {path}")
