"""
generate_report.py
------------------
Gera planilha Excel completa de todas as operações dos bots a partir do PostgreSQL.

Uso:
    python generate_report.py

Saída:
    D:/3 - Projetos investimentos/bot_taxa_cripto/relatorio_operacoes.xlsx
"""

from __future__ import annotations

import asyncio
import sys
from datetime import datetime, timezone, timedelta
from decimal import Decimal
from pathlib import Path

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Configuração
# ---------------------------------------------------------------------------

DB_URL = "postgres://vorxia:91318244@69.62.92.189:5432/vorxia"
OUTPUT_PATH = Path("D:/3 - Projetos investimentos/bot_taxa_cripto/relatorio_operacoes.xlsx")

BRT = timezone(timedelta(hours=-3))

# Cores
COR_HEADER_FILL = "1F3864"
COR_HEADER_FONT = "FFFFFF"
COR_LINHA_PAR = "EBF3FB"
COR_LINHA_IMPAR = "FFFFFF"
COR_PNL_POS_FILL = "C6EFCE"
COR_PNL_POS_FONT = "276221"
COR_PNL_NEG_FILL = "FFC7CE"
COR_PNL_NEG_FONT = "9C0006"
COR_RODAPE_FILL = "D9E1F2"


# ---------------------------------------------------------------------------
# Helpers de formatação
# ---------------------------------------------------------------------------

def to_brt(value: object) -> str:
    """Converte datetime (ou string ISO) para string BRT formatada."""
    if value is None:
        return ""
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return ""
        # Tenta parsear como ISO (pode ter ou não 'Z')
        for fmt in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z",
                    "%Y-%m-%d %H:%M:%S.%f%z", "%Y-%m-%d %H:%M:%S%z",
                    "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S"):
            try:
                dt = datetime.strptime(value, fmt)
                break
            except ValueError:
                continue
        else:
            return value  # devolve a string original se não parsear
    elif isinstance(value, datetime):
        dt = value
    else:
        return str(value)

    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    dt_brt = dt.astimezone(BRT)
    return dt_brt.strftime("%d/%m/%Y %H:%M:%S")


def parse_dt(value: object) -> datetime | None:
    """Converte valor em datetime com tz UTC (para cálculo de duração)."""
    if value is None:
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        for fmt in ("%Y-%m-%dT%H:%M:%S.%f%z", "%Y-%m-%dT%H:%M:%S%z",
                    "%Y-%m-%d %H:%M:%S.%f%z", "%Y-%m-%d %H:%M:%S%z",
                    "%Y-%m-%dT%H:%M:%S.%f", "%Y-%m-%dT%H:%M:%S",
                    "%Y-%m-%d %H:%M:%S"):
            try:
                dt = datetime.strptime(value, fmt)
                if dt.tzinfo is None:
                    dt = dt.replace(tzinfo=timezone.utc)
                return dt
            except ValueError:
                continue
    return None


def duracao_minutos(open_val: object, close_val: object) -> float | str:
    """Retorna duração em minutos (float) ou '' se não calculável."""
    dt_open = parse_dt(open_val)
    dt_close = parse_dt(close_val)
    if dt_open and dt_close and dt_close >= dt_open:
        delta = dt_close - dt_open
        return round(delta.total_seconds() / 60, 1)
    return ""


def to_float(value: object) -> float | None:
    """Converte Decimal/str/int/float para float, ou None."""
    if value is None:
        return None
    if isinstance(value, Decimal):
        return float(value)
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def profit_factor(pnls: list[float]) -> str:
    """Calcula Profit Factor = soma_ganhos / abs(soma_perdas)."""
    gross_profit = sum(p for p in pnls if p > 0)
    gross_loss = abs(sum(p for p in pnls if p < 0))
    if gross_loss == 0:
        return "inf" if gross_profit > 0 else "0"
    return f"{gross_profit / gross_loss:.4f}"


# ---------------------------------------------------------------------------
# Estilos openpyxl
# ---------------------------------------------------------------------------

def header_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COR_HEADER_FILL)


def header_font() -> Font:
    return Font(bold=True, color=COR_HEADER_FONT, name="Calibri", size=11)


def row_fill(idx: int) -> PatternFill:
    color = COR_LINHA_PAR if idx % 2 == 0 else COR_LINHA_IMPAR
    return PatternFill("solid", fgColor=color)


def pnl_fill(value: float | None) -> PatternFill | None:
    if value is None:
        return None
    if value > 0:
        return PatternFill("solid", fgColor=COR_PNL_POS_FILL)
    if value < 0:
        return PatternFill("solid", fgColor=COR_PNL_NEG_FILL)
    return None


def pnl_font(value: float | None) -> Font | None:
    if value is None:
        return None
    if value > 0:
        return Font(color=COR_PNL_POS_FONT, name="Calibri", size=10)
    if value < 0:
        return Font(color=COR_PNL_NEG_FONT, name="Calibri", size=10)
    return None


def rodape_fill() -> PatternFill:
    return PatternFill("solid", fgColor=COR_RODAPE_FILL)


def thin_border() -> Border:
    s = Side(border_style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def center_align() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def left_align() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def auto_largura(ws, min_w: int = 12, max_w: int = 40) -> None:
    """Ajusta largura de cada coluna baseado no conteúdo."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value) if cell.value is not None else ""
                max_len = max(max_len, len(val))
            except Exception:
                pass
        adjusted = max(min_w, min(max_len + 2, max_w))
        ws.column_dimensions[col_letter].width = adjusted


def aplicar_cabecalho(ws, colunas: list[str]) -> None:
    """Escreve e formata a linha de cabeçalho."""
    ws.append(colunas)
    for cell in ws[1]:
        cell.fill = header_fill()
        cell.font = header_font()
        cell.alignment = center_align()
        cell.border = thin_border()
    ws.row_dimensions[1].height = 22


def adicionar_rodape(ws, gerado_em: str) -> None:
    """Adiciona linha de rodapé com timestamp de geração."""
    ws.append([])
    rodape_row = ws.max_row + 1
    ws.append([f"Relatório gerado em {gerado_em} (BRT)"])
    cell = ws.cell(row=rodape_row, column=1)
    cell.fill = rodape_fill()
    cell.font = Font(italic=True, color="3F4F6B", name="Calibri", size=9)
    cell.alignment = left_align()


# ---------------------------------------------------------------------------
# Queries
# ---------------------------------------------------------------------------

SQL_OPERACOES = """
SELECT
    rt.id,
    rt.config_id,
    rc.session_name,
    rt.symbol,
    rt.direction,
    rt.entry_price,
    rt.exit_price,
    rt.total_pnl,
    rt.total_pnl_pct,
    rt.price_pnl,
    rt.funding_pnl,
    rt.fee_cost,
    rt.close_reason,
    rt.entry_score,
    rt.open_time,
    rt.close_time,
    rt.trade_timestamp,
    rt.exchange,
    rt.created_at
FROM real_trades rt
JOIN real_config rc ON rc.id = rt.config_id
WHERE rc.user_id = 1
ORDER BY rt.trade_timestamp DESC
"""

SQL_RESUMO_BOT = """
SELECT
    rc.id,
    rc.session_name,
    COUNT(*) AS total_trades,
    SUM(CASE WHEN rt.total_pnl > 0 THEN 1 ELSE 0 END) AS wins,
    SUM(CASE WHEN rt.total_pnl <= 0 THEN 1 ELSE 0 END) AS losses,
    ROUND(SUM(rt.total_pnl)::numeric, 4) AS pnl_total,
    ROUND(AVG(rt.total_pnl)::numeric, 4) AS avg_pnl,
    ROUND(MAX(rt.total_pnl)::numeric, 4) AS max_win,
    ROUND(MIN(rt.total_pnl)::numeric, 4) AS max_loss,
    ROUND(SUM(rt.fee_cost)::numeric, 4) AS total_fees,
    rc.capital
FROM real_trades rt
JOIN real_config rc ON rc.id = rt.config_id
WHERE rc.user_id = 1
GROUP BY rc.id, rc.session_name, rc.capital
ORDER BY pnl_total DESC
"""

SQL_RESUMO_BOT_PNLS = """
SELECT
    rt.config_id,
    rt.total_pnl
FROM real_trades rt
JOIN real_config rc ON rc.id = rt.config_id
WHERE rc.user_id = 1
ORDER BY rt.config_id
"""

SQL_RESUMO_SIMBOLO = """
SELECT
    rt.symbol,
    COUNT(*) AS total_trades,
    SUM(CASE WHEN rt.total_pnl > 0 THEN 1 ELSE 0 END) AS wins,
    SUM(CASE WHEN rt.total_pnl <= 0 THEN 1 ELSE 0 END) AS losses,
    ROUND(SUM(rt.total_pnl)::numeric, 4) AS pnl_total,
    ROUND(AVG(rt.total_pnl)::numeric, 4) AS avg_pnl,
    ROUND(SUM(CASE WHEN rt.total_pnl > 0 THEN rt.total_pnl ELSE 0 END)::numeric, 4) AS gross_profit,
    ROUND(ABS(SUM(CASE WHEN rt.total_pnl < 0 THEN rt.total_pnl ELSE 0 END))::numeric, 4) AS gross_loss
FROM real_trades rt
JOIN real_config rc ON rc.id = rt.config_id
WHERE rc.user_id = 1
GROUP BY rt.symbol
HAVING COUNT(*) >= 2
ORDER BY pnl_total DESC
"""


# ---------------------------------------------------------------------------
# Construção das abas
# ---------------------------------------------------------------------------

def build_aba_operacoes(ws, rows: list[asyncpg.Record]) -> None:
    """Monta a aba 'Todas as Operações'."""
    colunas = [
        "ID", "Bot ID", "Nome do Bot", "Símbolo", "Direção",
        "Preço Entrada", "Preço Saída", "PnL (USD)", "PnL (%)",
        "PnL Preço (USD)", "PnL Funding (USD)", "Taxa (USD)",
        "Motivo Fechamento", "Score de Entrada",
        "Data Abertura (BRT)", "Data Fechamento (BRT)",
        "Duração (min)", "Exchange",
    ]
    aplicar_cabecalho(ws, colunas)

    # Índices das colunas relevantes (1-based)
    COL_PNL_USD = 8
    COL_PRECO_ENTRADA = 6
    COL_PRECO_SAIDA = 7
    COL_PNL_PRECO = 10
    COL_PNL_FUNDING = 11
    COL_TAXA = 12

    fmt_preco = "#,##0.000000"
    fmt_pnl = '$#,##0.0000;[Red]-$#,##0.0000'

    for idx, row in enumerate(rows, start=2):
        # Definir abertura: prioridade open_time, fallback created_at
        open_raw = row["open_time"] or row["created_at"]
        close_raw = row["close_time"]

        pnl = to_float(row["total_pnl"])
        pnl_pct = to_float(row["total_pnl_pct"])
        pnl_pct_str = f"{pnl_pct:.4f}%" if pnl_pct is not None else ""

        linha = [
            row["id"],
            row["config_id"],
            row["session_name"] or "",
            row["symbol"] or "",
            row["direction"] or "",
            to_float(row["entry_price"]),
            to_float(row["exit_price"]),
            pnl,
            pnl_pct_str,
            to_float(row["price_pnl"]),
            to_float(row["funding_pnl"]),
            to_float(row["fee_cost"]),
            row["close_reason"] or "",
            to_float(row["entry_score"]),
            to_brt(open_raw),
            to_brt(close_raw),
            duracao_minutos(open_raw, close_raw),
            row["exchange"] or "",
        ]
        ws.append(linha)

        # Fundo alternado na linha
        fill = row_fill(idx)
        for cell in ws[idx]:
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = left_align()

        # Colorir coluna PnL USD
        pnl_cell = ws.cell(row=idx, column=COL_PNL_USD)
        f = pnl_fill(pnl)
        fnt = pnl_font(pnl)
        if f:
            pnl_cell.fill = f
        if fnt:
            pnl_cell.font = fnt
        pnl_cell.number_format = fmt_pnl

        # Formatos numéricos
        for col in (COL_PRECO_ENTRADA, COL_PRECO_SAIDA, COL_PNL_PRECO, COL_PNL_FUNDING):
            ws.cell(row=idx, column=col).number_format = fmt_preco
        ws.cell(row=idx, column=COL_TAXA).number_format = fmt_pnl

    ws.freeze_panes = "A2"
    auto_largura(ws)


def build_aba_resumo_bot(
    ws,
    rows: list[asyncpg.Record],
    pnls_por_bot: dict[int, list[float]],
) -> None:
    """Monta a aba 'Resumo por Bot'."""
    colunas = [
        "Bot ID", "Nome do Bot", "Total Trades", "Wins", "Losses",
        "Win Rate (%)", "PnL Total (USD)", "PnL Médio (USD)",
        "Maior Win (USD)", "Maior Loss (USD)", "Total Fees (USD)",
        "Profit Factor", "Capital",
    ]
    aplicar_cabecalho(ws, colunas)

    fmt_pnl = '$#,##0.0000;[Red]-$#,##0.0000'
    COL_PNL_TOTAL = 7

    for idx, row in enumerate(rows, start=2):
        config_id: int = row["id"]
        total: int = int(row["total_trades"])
        wins: int = int(row["wins"])
        win_rate = round(wins / total * 100, 2) if total > 0 else 0.0
        pf = profit_factor(pnls_por_bot.get(config_id, []))
        pnl = to_float(row["pnl_total"])

        linha = [
            config_id,
            row["session_name"] or "",
            total,
            wins,
            int(row["losses"]),
            win_rate,
            pnl,
            to_float(row["avg_pnl"]),
            to_float(row["max_win"]),
            to_float(row["max_loss"]),
            to_float(row["total_fees"]),
            pf,
            to_float(row["capital"]),
        ]
        ws.append(linha)

        fill = row_fill(idx)
        for cell in ws[idx]:
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = left_align()

        # Colorir PnL Total
        pnl_cell = ws.cell(row=idx, column=COL_PNL_TOTAL)
        f = pnl_fill(pnl)
        fnt = pnl_font(pnl)
        if f:
            pnl_cell.fill = f
        if fnt:
            pnl_cell.font = fnt

        # Formatos numéricos
        for col in (COL_PNL_TOTAL, 8, 9, 10, 11, 13):
            ws.cell(row=idx, column=col).number_format = fmt_pnl
        ws.cell(row=idx, column=6).number_format = '0.00"%"'

    ws.freeze_panes = "A2"
    auto_largura(ws)


def build_aba_resumo_simbolo(ws, rows: list[asyncpg.Record]) -> None:
    """Monta a aba 'Resumo por Símbolo'."""
    colunas = [
        "Símbolo", "Total Trades", "Wins", "Losses",
        "Win Rate (%)", "PnL Total (USD)", "PnL Médio (USD)",
        "Profit Factor",
    ]
    aplicar_cabecalho(ws, colunas)

    fmt_pnl = '$#,##0.0000;[Red]-$#,##0.0000'
    COL_PNL_TOTAL = 6

    for idx, row in enumerate(rows, start=2):
        total: int = int(row["total_trades"])
        wins: int = int(row["wins"])
        win_rate = round(wins / total * 100, 2) if total > 0 else 0.0
        gross_profit = to_float(row["gross_profit"]) or 0.0
        gross_loss = to_float(row["gross_loss"]) or 0.0
        pf = f"{gross_profit / gross_loss:.4f}" if gross_loss > 0 else ("inf" if gross_profit > 0 else "0")
        pnl = to_float(row["pnl_total"])

        linha = [
            row["symbol"] or "",
            total,
            wins,
            int(row["losses"]),
            win_rate,
            pnl,
            to_float(row["avg_pnl"]),
            pf,
        ]
        ws.append(linha)

        fill = row_fill(idx)
        for cell in ws[idx]:
            cell.fill = fill
            cell.border = thin_border()
            cell.alignment = left_align()

        pnl_cell = ws.cell(row=idx, column=COL_PNL_TOTAL)
        f = pnl_fill(pnl)
        fnt = pnl_font(pnl)
        if f:
            pnl_cell.fill = f
        if fnt:
            pnl_cell.font = fnt

        for col in (COL_PNL_TOTAL, 7):
            ws.cell(row=idx, column=col).number_format = fmt_pnl
        ws.cell(row=idx, column=5).number_format = '0.00"%"'

    ws.freeze_panes = "A2"
    auto_largura(ws)


# ---------------------------------------------------------------------------
# Main async
# ---------------------------------------------------------------------------

async def gerar_relatorio() -> None:
    print("[1/6] Conectando ao banco de dados...")
    conn = await asyncpg.connect(DB_URL)

    try:
        print("[2/6] Buscando operações...")
        rows_op = await conn.fetch(SQL_OPERACOES)
        print(f"      {len(rows_op)} operações encontradas.")

        print("[3/6] Buscando resumo por bot...")
        rows_bot = await conn.fetch(SQL_RESUMO_BOT)

        print("[4/6] Buscando PnLs por bot para Profit Factor...")
        rows_pnls = await conn.fetch(SQL_RESUMO_BOT_PNLS)

        print("[5/6] Buscando resumo por símbolo...")
        rows_sim = await conn.fetch(SQL_RESUMO_SIMBOLO)
    finally:
        await conn.close()

    # Agrupa PnLs por config_id para cálculo do Profit Factor
    pnls_por_bot: dict[int, list[float]] = {}
    for r in rows_pnls:
        cid: int = r["config_id"]
        pnl = to_float(r["total_pnl"])
        if pnl is not None:
            pnls_por_bot.setdefault(cid, []).append(pnl)

    print("[6/6] Gerando planilha Excel...")
    wb = Workbook()

    # Remove a aba padrão criada pelo openpyxl
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    gerado_em = datetime.now(BRT).strftime("%d/%m/%Y %H:%M:%S")

    # Aba 1
    ws1 = wb.create_sheet("Todas as Operações")
    ws1.sheet_view.showGridLines = True
    build_aba_operacoes(ws1, list(rows_op))
    adicionar_rodape(ws1, gerado_em)
    print(f"      Aba 'Todas as Operações': {len(rows_op)} linhas.")

    # Aba 2
    ws2 = wb.create_sheet("Resumo por Bot")
    ws2.sheet_view.showGridLines = True
    build_aba_resumo_bot(ws2, list(rows_bot), pnls_por_bot)
    adicionar_rodape(ws2, gerado_em)
    print(f"      Aba 'Resumo por Bot': {len(rows_bot)} bots.")

    # Aba 3
    ws3 = wb.create_sheet("Resumo por Símbolo")
    ws3.sheet_view.showGridLines = True
    build_aba_resumo_simbolo(ws3, list(rows_sim))
    adicionar_rodape(ws3, gerado_em)
    print(f"      Aba 'Resumo por Símbolo': {len(rows_sim)} símbolos.")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(OUTPUT_PATH))
    print(f"\nPlanilha salva em: {OUTPUT_PATH}")
    print(f"Gerado em: {gerado_em} (BRT)")


def main() -> None:
    try:
        asyncio.run(gerar_relatorio())
    except KeyboardInterrupt:
        print("\nInterrompido pelo usuário.")
        sys.exit(1)
    except Exception as exc:
        print(f"\nERRO: {exc}", file=sys.stderr)
        raise


if __name__ == "__main__":
    main()
